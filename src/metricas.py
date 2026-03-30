"""
metricas.py — Cálculo de métricas clave de quejas
Punto b de la prueba técnica
"""
import pandas as pd
from collections import Counter


def load_data(path: str = r"./data/BD_Quejas_Analitica.xlsx") -> pd.DataFrame:
    df = pd.read_excel(path)
    df["Mes"] = df["Mes apertura del caso"].astype(str)
    df["Categoria"] = df["Descripción"].apply(classify_queja)
    return df


def classify_queja(text: str) -> str:
    text = str(text).upper()
    if ("PAGO" in text or "COBRO" in text or "MORA" in text or
            "SALARIO" in text or "PLATA" in text) and "INCAPACIDAD" in text:
        return "DEMORA_PAGO_INCAPACIDAD"
    if "INCAPACIDAD" in text and any(w in text for w in
                                      ["ESTADO", "INFORMACIÓN", "INFORMACION",
                                       "CONSULTA", "AVERIGUAR", "SABER"]):
        return "CONSULTA_ESTADO_INCAPACIDAD"
    if "INCAPACIDAD" in text and "RADICADO" in text:
        return "SEGUIMIENTO_RADICADO"
    if "CERTIFICADO" in text:
        return "SOLICITUD_CERTIFICADO"
    if "ACCIDENTE" in text and "TRABAJO" in text:
        return "ACCIDENTE_TRABAJO"
    if "INCAPACIDAD" in text:
        return "GESTION_INCAPACIDAD"
    return "OTRA_SOLICITUD"


def calcular_metricas(df: pd.DataFrame) -> dict:
    total = len(df)

    # 1. Volumen total de quejas
    volumen_total = total

    # 2. Quejas por mes
    quejas_por_mes = df.groupby("Mes").size().to_dict()

    # 3. Distribución por canal
    canal_dist = df["Canal de comunicación"].value_counts().to_dict()
    canal_pct = (df["Canal de comunicación"].value_counts(normalize=True) * 100).round(2).to_dict()

    # 4. Distribución por categoría de queja
    cat_dist = df["Categoria"].value_counts().to_dict()
    cat_pct = (df["Categoria"].value_counts(normalize=True) * 100).round(2).to_dict()

    # 5. Clientes recurrentes (más de 1 queja)
    clientes_count = df["Nombre del cliente"].value_counts()
    clientes_recurrentes = clientes_count[clientes_count > 1]
    tasa_recurrencia = round(len(clientes_recurrentes) / df["Nombre del cliente"].nunique() * 100, 2)

    # 6. Top 10 clientes con más quejas
    top_clientes = clientes_count.head(10).to_dict()

    # 7. Canal con mayor crecimiento (mes a mes)
    canal_mes = df.groupby(["Mes", "Canal de comunicación"]).size().unstack(fill_value=0)

    # 8. Categoría dominante
    categoria_dominante = df["Categoria"].value_counts().idxmax()
    pct_dominante = round(df["Categoria"].value_counts(normalize=True).max() * 100, 2)

    # 9. Quejas por canal y categoría (cruce)
    cruce_canal_cat = df.groupby(["Canal de comunicación", "Categoria"]).size().unstack(fill_value=0).to_dict()

    return {
        "volumen_total": volumen_total,
        "quejas_por_mes": quejas_por_mes,
        "canal_distribucion": canal_dist,
        "canal_porcentaje": canal_pct,
        "categoria_distribucion": cat_dist,
        "categoria_porcentaje": cat_pct,
        "clientes_unicos": df["Nombre del cliente"].nunique(),
        "clientes_recurrentes": len(clientes_recurrentes),
        "tasa_recurrencia_pct": tasa_recurrencia,
        "top_10_clientes": top_clientes,
        "categoria_dominante": categoria_dominante,
        "pct_categoria_dominante": pct_dominante,
    }


def exportar_metricas(metricas: dict, df: pd.DataFrame, out_path: str = r"./outputs/metricas_resumen.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Hoja 1: Resumen general
    ws1 = wb.active
    ws1.title = "Resumen General"
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    resumen = [
        ["Métrica", "Valor"],
        ["Total Quejas", metricas["volumen_total"]],
        ["Clientes Únicos", metricas["clientes_unicos"]],
        ["Clientes Recurrentes", metricas["clientes_recurrentes"]],
        ["Tasa de Recurrencia (%)", metricas["tasa_recurrencia_pct"]],
        ["Categoría Dominante", metricas["categoria_dominante"]],
        ["% Categoría Dominante", metricas["pct_categoria_dominante"]],
    ]
    for r, row in enumerate(resumen, 1):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill

    # Hoja 2: Quejas por mes
    ws2 = wb.create_sheet("Quejas por Mes")
    ws2.append(["Mes", "Total Quejas"])
    for mes, cnt in sorted(metricas["quejas_por_mes"].items()):
        ws2.append([mes, cnt])

    # Hoja 3: Por canal
    ws3 = wb.create_sheet("Por Canal")
    ws3.append(["Canal", "Total", "Porcentaje (%)"])
    for canal, cnt in metricas["canal_distribucion"].items():
        pct = metricas["canal_porcentaje"].get(canal, 0)
        ws3.append([canal, cnt, pct])

    # Hoja 4: Por categoría
    ws4 = wb.create_sheet("Por Categoría")
    ws4.append(["Categoría", "Total", "Porcentaje (%)"])
    for cat, cnt in metricas["categoria_distribucion"].items():
        pct = metricas["categoria_porcentaje"].get(cat, 0)
        ws4.append([cat, cnt, pct])

    # Hoja 5: Top clientes
    ws5 = wb.create_sheet("Top Clientes")
    ws5.append(["Cliente", "Nº Quejas"])
    for cliente, cnt in metricas["top_10_clientes"].items():
        ws5.append([cliente, cnt])

    wb.save(out_path)
    print(f"✅ Métricas exportadas a {out_path}")


if __name__ == "__main__":
    df = load_data(r"./data/BD_Quejas_Analitica.xlsx")
    metricas = calcular_metricas(df)
    print("\n📊 MÉTRICAS CLAVE")
    print("=" * 50)
    for k, v in metricas.items():
        if not isinstance(v, dict):
            print(f"  {k}: {v}")
    exportar_metricas(metricas, df)
